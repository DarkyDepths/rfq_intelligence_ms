"""Compliance and deviation extraction for package parser Stage C."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from src.services.package_parser.contracts import (
    ComplianceLineItem,
    ComplianceProfile,
    DeviationProfile,
    FileEntry,
    PackageInventory,
    SectionMatch,
    SectionRegistry,
)


_HEADER_NORMALIZE_RE = re.compile(r"[^a-z0-9]+")
_ITEM_NO_RE = re.compile(r"^\d+\.\d+$")
_COMPLIANCE_METADATA_ALIASES: dict[str, tuple[str, ...]] = {
    "material_description": ("materialdescription", "materialname", "materialtitle"),
    "mr_number": ("mrnumber", "mrno"),
    "nine_com": ("9com", "ninecom", "9comcode"),
}
_DEVIATION_METADATA_ALIASES: dict[str, tuple[str, ...]] = {
    "bi_number": ("binumber", "bino"),
    "project_title": ("projecttitle", "projectname"),
    "mr_number": ("mrnumber", "mrno"),
    "material_title": ("materialtitle", "materialdescription", "materialname"),
}
_COMPLIANCE_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "item_no": ("itemno", "item", "clause"),
    "description": ("description",),
    "specified_requirement": ("specifiedrequirement", "requirement"),
    "section_label": ("sectionlabel", "section"),
}


class ComplianceExtractor:
    """Extract compliance sheet items and deviation metadata from folder 15."""

    def extract(
        self,
        inventory: PackageInventory,
        registry: SectionRegistry,
        package_root: Path,
    ) -> tuple[ComplianceProfile | None, DeviationProfile | None]:
        section = self._find_section(registry)
        if section is None:
            return None, None

        compliance_file = self._find_workbook(
            inventory=inventory,
            section=section,
            required_terms=("compliance", "sheet"),
        )
        deviation_file = self._find_workbook(
            inventory=inventory,
            section=section,
            required_terms=("deviation", "list"),
        )

        compliance_profile = self._extract_compliance_profile(package_root, compliance_file)
        deviation_profile = self._extract_deviation_profile(package_root, deviation_file)
        return compliance_profile, deviation_profile

    @staticmethod
    def _find_section(registry: SectionRegistry) -> SectionMatch | None:
        for section in registry.matched_sections:
            if section.canonical_key == "notes_to_vendor":
                return section
        return None

    def _find_workbook(
        self,
        inventory: PackageInventory,
        section: SectionMatch,
        required_terms: tuple[str, ...],
    ) -> FileEntry | None:
        section_prefix = f"{section.folder_relative_path}/"
        candidates = [
            file_entry
            for file_entry in inventory.files
            if not file_entry.is_system_file
            and not file_entry.is_mr_index
            and file_entry.extension.lower() == ".xlsx"
            and file_entry.relative_path.startswith(section_prefix)
            and all(term in file_entry.filename.lower() for term in required_terms)
        ]
        if not candidates:
            return None
        return sorted(candidates, key=lambda item: item.relative_path.lower())[0]

    def _extract_compliance_profile(
        self,
        package_root: Path,
        source_file: FileEntry | None,
    ) -> ComplianceProfile | None:
        if source_file is None:
            return None

        workbook = self._load_workbook(package_root, source_file)
        if workbook is None:
            return None

        try:
            worksheet = workbook.worksheets[0]
            metadata = self._extract_labeled_metadata(worksheet, _COMPLIANCE_METADATA_ALIASES, max_row=12, max_col=12)
            metadata.setdefault("material_description", self._cell_to_text(worksheet["I1"].value))
            metadata.setdefault("mr_number", self._cell_to_text(worksheet["I3"].value))
            metadata.setdefault("nine_com", self._cell_to_text(worksheet["I4"].value))

            line_items = self._parse_compliance_line_items(worksheet)
            return ComplianceProfile(
                source_file=source_file.relative_path,
                line_items=line_items,
                total_items=len(line_items),
                section_labels=self._ordered_unique(item.section_label for item in line_items),
                material_description=metadata.get("material_description"),
                mr_number=metadata.get("mr_number"),
                nine_com=metadata.get("nine_com"),
            )
        finally:
            workbook.close()

    def _extract_deviation_profile(
        self,
        package_root: Path,
        source_file: FileEntry | None,
    ) -> DeviationProfile | None:
        if source_file is None:
            return None

        workbook = self._load_workbook(package_root, source_file)
        if workbook is None:
            return None

        try:
            worksheet = workbook.worksheets[0]
            metadata = self._extract_labeled_metadata(worksheet, _DEVIATION_METADATA_ALIASES, max_row=15, max_col=10)
            metadata.setdefault("mr_number", self._cell_to_text(worksheet["D8"].value))

            data_start_row = self._find_deviation_data_start_row(worksheet)
            total_rows = 0
            has_vendor_entries = False
            for row_values in worksheet.iter_rows(min_row=data_start_row, values_only=True):
                non_empty = [self._cell_to_text(value) for value in row_values if self._cell_to_text(value) is not None]
                if not non_empty:
                    continue
                total_rows += 1
                if len(non_empty) >= 2:
                    has_vendor_entries = True

            return DeviationProfile(
                source_file=source_file.relative_path,
                total_rows=total_rows,
                has_vendor_entries=has_vendor_entries,
                bi_number=metadata.get("bi_number"),
                project_title=metadata.get("project_title"),
                mr_number=metadata.get("mr_number"),
                material_title=metadata.get("material_title"),
            )
        finally:
            workbook.close()

    def _load_workbook(self, package_root: Path, source_file: FileEntry):
        try:
            return load_workbook(package_root / source_file.relative_path, read_only=True, data_only=True)
        except (FileNotFoundError, InvalidFileException, OSError):
            return None

    def _extract_labeled_metadata(
        self,
        worksheet,
        aliases_map: dict[str, tuple[str, ...]],
        max_row: int,
        max_col: int,
    ) -> dict[str, str | None]:
        results: dict[str, str | None] = {}
        scan_row_limit = min(worksheet.max_row or 0, max_row)
        scan_col_limit = min(worksheet.max_column or 0, max_col)

        for row_idx in range(1, scan_row_limit + 1):
            for col_idx in range(1, scan_col_limit + 1):
                normalized = self._normalize_label(worksheet.cell(row=row_idx, column=col_idx).value)
                if normalized is None:
                    continue

                for field_name, aliases in aliases_map.items():
                    if field_name in results or normalized not in aliases:
                        continue
                    results[field_name] = self._find_adjacent_metadata_value(
                        worksheet=worksheet,
                        row_idx=row_idx,
                        col_idx=col_idx,
                        max_row=scan_row_limit,
                        max_col=scan_col_limit,
                    )
                    break

        return results

    def _find_adjacent_metadata_value(
        self,
        worksheet,
        row_idx: int,
        col_idx: int,
        max_row: int,
        max_col: int,
    ) -> str | None:
        for look_col in range(col_idx + 1, max_col + 1):
            value = self._cell_to_text(worksheet.cell(row=row_idx, column=look_col).value)
            if value is not None:
                return value

        for look_row in range(row_idx + 1, min(max_row, row_idx + 2) + 1):
            value = self._cell_to_text(worksheet.cell(row=look_row, column=col_idx).value)
            if value is not None:
                return value

        return None

    def _parse_compliance_line_items(self, worksheet) -> list[ComplianceLineItem]:
        header_row, column_map = self._find_compliance_header_row(worksheet)
        if column_map is None:
            item_col = 2
            description_col = 4
            requirement_col = 7
            start_row = 1
        else:
            item_col = column_map.get("item_no", 2)
            description_col = column_map.get("description", 4)
            requirement_col = column_map.get("specified_requirement", 7)
            start_row = header_row + 1

        current_section_label: str | None = None
        line_items: list[ComplianceLineItem] = []

        for row_idx, row_values in enumerate(
            worksheet.iter_rows(min_row=start_row, values_only=True),
            start=start_row,
        ):
            item_no = self._item_text(self._value_at(row_values, item_col))
            if item_no is None or _ITEM_NO_RE.fullmatch(item_no) is None:
                continue

            description = self._cell_to_text(self._value_at(row_values, description_col))
            requirement = self._cell_to_text(self._value_at(row_values, requirement_col))

            if item_no.endswith(".0"):
                current_section_label = description or self._first_nonempty_text(row_values, skip_cols={item_col})
                continue

            if description is None and requirement is None:
                continue

            line_items.append(
                ComplianceLineItem(
                    sheet_row=row_idx,
                    item_no=item_no,
                    description=description,
                    specified_requirement=requirement,
                    section_label=current_section_label,
                )
            )

        return line_items

    def _find_compliance_header_row(self, worksheet) -> tuple[int, dict[str, int] | None]:
        scan_limit = min(worksheet.max_row or 0, 20)
        for row_idx, row_values in enumerate(
            worksheet.iter_rows(min_row=1, max_row=scan_limit, values_only=True),
            start=1,
        ):
            column_map = self._build_column_map(row_values, _COMPLIANCE_HEADER_ALIASES)
            if "item_no" in column_map and ("description" in column_map or "specified_requirement" in column_map):
                return row_idx, column_map
        return 0, None

    def _find_deviation_data_start_row(self, worksheet) -> int:
        scan_limit = min(worksheet.max_row or 0, 25)
        for row_idx, row_values in enumerate(
            worksheet.iter_rows(min_row=1, max_row=scan_limit, values_only=True),
            start=1,
        ):
            normalized_values = {self._normalize_label(value) for value in row_values}
            if {"vendor", "description"} & {value for value in normalized_values if value is not None}:
                return row_idx + 1
        return min(11, (worksheet.max_row or 1) + 1)

    def _build_column_map(
        self,
        row_values: Iterable[object],
        aliases_map: dict[str, tuple[str, ...]],
    ) -> dict[str, int]:
        column_map: dict[str, int] = {}
        for column_idx, cell_value in enumerate(row_values, start=1):
            normalized = self._normalize_label(cell_value)
            if normalized is None:
                continue
            for field_name, aliases in aliases_map.items():
                if field_name in column_map or normalized not in aliases:
                    continue
                column_map[field_name] = column_idx
                break
        return column_map

    @staticmethod
    def _value_at(row_values: tuple[object, ...], column_index: int) -> object | None:
        if column_index < 1 or column_index > len(row_values):
            return None
        return row_values[column_index - 1]

    @staticmethod
    def _normalize_label(value: object) -> str | None:
        text = ComplianceExtractor._cell_to_text(value)
        if text is None:
            return None
        normalized = _HEADER_NORMALIZE_RE.sub("", text.lower())
        return normalized or None

    @staticmethod
    def _cell_to_text(value: object) -> str | None:
        if value is None:
            return None
        if isinstance(value, str):
            text = " ".join(value.split())
            return text or None
        if isinstance(value, bool):
            return "TRUE" if value else "FALSE"
        if isinstance(value, int):
            return str(value)
        if isinstance(value, float):
            return str(int(value)) if value.is_integer() else str(value)
        text = str(value).strip()
        return text or None

    @staticmethod
    def _item_text(value: object) -> str | None:
        if value is None:
            return None
        if isinstance(value, bool):
            return None
        if isinstance(value, int):
            return f"{value}.0"
        if isinstance(value, float):
            return f"{value:.1f}" if value.is_integer() else str(value)
        text = " ".join(str(value).split())
        return text or None

    def _first_nonempty_text(self, row_values: tuple[object, ...], skip_cols: set[int]) -> str | None:
        for idx, value in enumerate(row_values, start=1):
            if idx in skip_cols:
                continue
            text = self._cell_to_text(value)
            if text is not None:
                return text
        return None

    @staticmethod
    def _ordered_unique(values: Iterable[str | None]) -> list[str]:
        seen: set[str] = set()
        ordered: list[str] = []
        for value in values:
            if value is None or value in seen:
                continue
            seen.add(value)
            ordered.append(value)
        return ordered
