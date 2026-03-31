"""BOM extraction for package parser Stage C."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from src.services.package_parser.contracts import (
    BomLineItem,
    BomProfile,
    FileEntry,
    PackageInventory,
    SectionMatch,
    SectionRegistry,
)


_HEADER_SCAN_LIMIT = 25
_HEADER_NORMALIZE_RE = re.compile(r"[^a-z0-9]+")
_HEADER_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "mr_line_item": ("mrli", "mrlineitem"),
    "line_item": ("lineitem",),
    "nine_com": ("9com",),
    "plant_no": ("plantno", "plantnumber"),
    "pipeline": ("pipeline",),
    "design_code": ("designcode",),
    "service": ("service",),
    "material_type": ("matleqpttype", "materialtype", "matltype", "equipmenttype", "eqpttype"),
    "location": ("location",),
    "technical_spec": ("technicalspec", "technicalspecification", "techspec"),
    "tag_number": ("tagnumber", "tagno"),
    "data_sheet": ("datasheet",),
    "reference_drawings": ("referencedrawings", "engineeringreferencedrawings"),
    "quantity": ("qty", "quantity"),
}
_STRONG_HEADER_FIELDS = {
    "mr_line_item",
    "line_item",
    "nine_com",
    "plant_no",
    "design_code",
    "service",
    "location",
    "tag_number",
    "quantity",
}


class BomExtractor:
    """Extract BOM workbook contents from the canonical folder 02 section."""

    def extract(
        self,
        inventory: PackageInventory,
        registry: SectionRegistry,
        package_root: Path,
    ) -> BomProfile | None:
        section = self._find_bom_section(registry)
        if section is None:
            return None

        bom_file = self._find_bom_workbook(inventory, section)
        if bom_file is None:
            return None

        try:
            workbook = load_workbook(package_root / bom_file.relative_path, read_only=True, data_only=True)
        except FileNotFoundError:
            return None
        except InvalidFileException:
            return None
        except OSError:
            return None

        try:
            worksheet, header_row, column_map = self._find_target_sheet(workbook)
            if worksheet is None or header_row is None or column_map is None:
                return None

            line_items = self._parse_line_items(worksheet, header_row, column_map)
            return BomProfile(
                source_file=bom_file.relative_path,
                sheet_name=worksheet.title,
                line_items=line_items,
                total_line_items=len(line_items),
                tag_numbers_found=self._ordered_unique(item.tag_number for item in line_items),
                nine_com_codes_found=self._ordered_unique(item.nine_com for item in line_items),
                design_codes_found=self._ordered_unique(item.design_code for item in line_items),
                locations_found=self._ordered_unique(item.location for item in line_items),
            )
        finally:
            workbook.close()

    @staticmethod
    def _find_bom_section(registry: SectionRegistry) -> SectionMatch | None:
        for section in registry.matched_sections:
            if section.canonical_key == "description_bom":
                return section
        return None

    def _find_bom_workbook(self, inventory: PackageInventory, section: SectionMatch) -> FileEntry | None:
        section_prefix = f"{section.folder_relative_path}/"
        candidates = [
            file_entry
            for file_entry in inventory.files
            if not file_entry.is_system_file
            and not file_entry.is_mr_index
            and file_entry.extension.lower() == ".xlsx"
            and file_entry.relative_path.startswith(section_prefix)
            and "bom" in file_entry.filename.lower()
        ]
        if not candidates:
            return None
        return sorted(candidates, key=lambda item: item.relative_path.lower())[0]

    def _find_target_sheet(self, workbook) -> tuple[object | None, int | None, dict[str, int] | None]:
        preferred_sheets = [
            worksheet
            for worksheet in workbook.worksheets
            if worksheet.title.strip().lower() == "bom"
        ]
        fallback_sheets = [
            worksheet
            for worksheet in workbook.worksheets
            if worksheet.title.strip().lower() != "bom"
        ]

        for worksheet in [*preferred_sheets, *fallback_sheets]:
            header_row, column_map = self._find_header_row(worksheet)
            if header_row is not None and column_map is not None:
                return worksheet, header_row, column_map

        return None, None, None

    def _find_header_row(self, worksheet) -> tuple[int | None, dict[str, int] | None]:
        max_row = worksheet.max_row or 0
        scan_limit = min(max_row, _HEADER_SCAN_LIMIT)

        for row_idx, row_values in enumerate(
            worksheet.iter_rows(min_row=1, max_row=scan_limit, values_only=True),
            start=1,
        ):
            column_map = self._build_column_map(row_values)
            if self._is_header_map(column_map):
                return row_idx, column_map

        return None, None

    def _build_column_map(self, row_values: Iterable[object]) -> dict[str, int]:
        column_map: dict[str, int] = {}

        for column_idx, cell_value in enumerate(row_values, start=1):
            normalized = self._normalize_header(cell_value)
            if normalized is None:
                continue

            for field_name, aliases in _HEADER_FIELD_ALIASES.items():
                if field_name in column_map:
                    continue
                if normalized in aliases:
                    column_map[field_name] = column_idx
                    break

        return column_map

    @staticmethod
    def _is_header_map(column_map: dict[str, int]) -> bool:
        recognized_count = len(column_map)
        strong_count = sum(1 for field_name in column_map if field_name in _STRONG_HEADER_FIELDS)
        return recognized_count >= 4 and strong_count >= 3

    def _parse_line_items(self, worksheet, header_row: int, column_map: dict[str, int]) -> list[BomLineItem]:
        line_items: list[BomLineItem] = []

        for row_idx, row_values in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            line_item = BomLineItem(
                sheet_row=row_idx,
                mr_line_item=self._cell_to_text(self._value_at(row_values, column_map.get("mr_line_item"))),
                line_item=self._cell_to_text(self._value_at(row_values, column_map.get("line_item"))),
                nine_com=self._cell_to_text(self._value_at(row_values, column_map.get("nine_com"))),
                plant_no=self._cell_to_text(self._value_at(row_values, column_map.get("plant_no"))),
                pipeline=self._cell_to_text(self._value_at(row_values, column_map.get("pipeline"))),
                design_code=self._cell_to_text(self._value_at(row_values, column_map.get("design_code"))),
                service=self._cell_to_text(self._value_at(row_values, column_map.get("service"))),
                material_type=self._cell_to_text(self._value_at(row_values, column_map.get("material_type"))),
                location=self._cell_to_text(self._value_at(row_values, column_map.get("location"))),
                technical_spec=self._cell_to_text(self._value_at(row_values, column_map.get("technical_spec"))),
                tag_number=self._cell_to_text(self._value_at(row_values, column_map.get("tag_number"))),
                data_sheet=self._cell_to_text(self._value_at(row_values, column_map.get("data_sheet"))),
                reference_drawings=self._cell_to_text(self._value_at(row_values, column_map.get("reference_drawings"))),
                quantity=self._to_float(self._value_at(row_values, column_map.get("quantity"))),
            )
            if self._is_meaningful_line_item(line_item):
                line_items.append(line_item)

        return line_items

    @staticmethod
    def _value_at(row_values: tuple[object, ...], column_index: int | None) -> object | None:
        if column_index is None or column_index < 1 or column_index > len(row_values):
            return None
        return row_values[column_index - 1]

    @staticmethod
    def _normalize_header(value: object) -> str | None:
        text = BomExtractor._cell_to_text(value)
        if text is None:
            return None
        normalized = _HEADER_NORMALIZE_RE.sub("", text.lower())
        return normalized or None

    @staticmethod
    def _cell_to_text(value: object) -> str | None:
        if value is None:
            return None
        if isinstance(value, str):
            text = " ".join(value.split())
            return text or None
        if isinstance(value, bool):
            return "TRUE" if value else "FALSE"
        if isinstance(value, int):
            return str(value)
        if isinstance(value, float):
            return str(int(value)) if value.is_integer() else str(value)
        text = str(value).strip()
        return text or None

    @staticmethod
    def _to_float(value: object) -> float | None:
        if value is None or value == "":
            return None
        if isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            return float(value)
        try:
            return float(str(value).strip())
        except ValueError:
            return None

    @staticmethod
    def _is_meaningful_line_item(line_item: BomLineItem) -> bool:
        anchor_fields = (
            line_item.mr_line_item,
            line_item.line_item,
            line_item.nine_com,
            line_item.plant_no,
            line_item.tag_number,
        )
        if any(value is not None for value in anchor_fields):
            return True

        populated_fields = sum(
            1
            for value in (
                line_item.pipeline,
                line_item.design_code,
                line_item.service,
                line_item.material_type,
                line_item.location,
                line_item.technical_spec,
                line_item.data_sheet,
                line_item.reference_drawings,
                line_item.quantity,
            )
            if value is not None
        )
        return populated_fields >= 3

    @staticmethod
    def _ordered_unique(values: Iterable[str | None]) -> list[str]:
        seen: set[str] = set()
        ordered: list[str] = []

        for value in values:
            if value is None or value in seen:
                continue
            seen.add(value)
            ordered.append(value)

        return ordered
